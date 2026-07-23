"""
Excel 工具层
-----------
支持 openpyxl（.xlsx 读写）和 xlwings（通过 COM 操作 WPS/Excel 应用程序）。
确定性脚本，不调 AI。
"""

import logging
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class ExcelTool:
    """Excel 读写工具。静态方法，无状态。"""

    # ------------------------------------------------------------------
    # 内部辅助
    # ------------------------------------------------------------------

    @staticmethod
    def _ensure_path(filepath: str) -> Path:
        """将字符串路径转为 Path 并校验扩展名。"""
        path = Path(filepath)
        if path.suffix.lower() not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            logger.warning("非标准 Excel 扩展名: %s，仍然尝试打开", path.suffix)
        return path

    @staticmethod
    def _get_sheet(wb: openpyxl.Workbook, sheet_name: str | None) -> Worksheet:
        """根据名称获取 sheet；None 时返回活动 sheet。"""
        if sheet_name is None:
            return wb.active
        try:
            return wb[sheet_name]
        except KeyError:
            available = wb.sheetnames
            raise ValueError(
                f"Sheet '{sheet_name}' 不存在。可用 sheet: {available}"
            ) from None

    @staticmethod
    def _col_letters(n: int) -> list[str]:
        """生成前 n 列字母：A, B, C, ..., Z, AA, AB, ..."""
        result = []
        for i in range(n):
            result.append(get_column_letter(i + 1))
        return result

    @staticmethod
    def _cell_to_value(cell) -> Any:
        """
        将 openpyxl 单元格对象转为 Python 原生值。
        保留 datetime / int / float / str / None / bool 等类型。
        """
        # openpyxl 单元格的值已经是转换后的 Python 对象（datetime/int/float/str/None/bool）
        # 对于公式单元格，如果 load_workbook 使用了 data_only=True，这里拿到的是缓存值
        return cell.value

    # ------------------------------------------------------------------
    # 公开 API
    # ------------------------------------------------------------------

    @staticmethod
    def read(
        filepath: str,
        sheet_name: str | None = None,
        has_header: bool = True,
    ) -> list[dict]:
        """
        读取 Excel 文件，返回字典列表。

        Parameters
        ----------
        filepath : str
            Excel 文件路径（.xlsx / .xlsm 等）。
        sheet_name : str | None
            要读取的 sheet 名称，None 表示读取第一个 sheet。
        has_header : bool
            True → 第一行作为列名（key）。
            False → 列名用 A, B, C...

        Returns
        -------
        list[dict]
            例如 [{"日期": "2026-07-23", "金额": 386}, ...]
        """
        filepath = str(filepath)
        path = ExcelTool._ensure_path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {filepath}")

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            ws = ExcelTool._get_sheet(wb, sheet_name)
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        if not rows:
            logger.info("文件 %s 的 sheet '%s' 为空", filepath, ws.title)
            return []

        if has_header:
            headers = [str(h) if h is not None else "" for h in rows[0]]
            data = []
            for row in rows[1:]:
                record = {}
                for i, header in enumerate(headers):
                    record[header] = row[i] if i < len(row) else None
                data.append(record)
            return data
        else:
            col_count = max(len(r) for r in rows) if rows else 0
            headers = ExcelTool._col_letters(col_count)
            data = []
            for row in rows:
                record = {}
                for i, header in enumerate(headers):
                    record[header] = row[i] if i < len(row) else None
                data.append(record)
            return data

    @staticmethod
    def read_range(
        filepath: str,
        sheet_name: str,
        start_cell: str,
        end_cell: str,
    ) -> list[list[Any]]:
        """
        读取指定区域。

        Parameters
        ----------
        filepath : str
            Excel 文件路径。
        sheet_name : str
            Sheet 名称。
        start_cell : str
            起始单元格，如 "A2"。
        end_cell : str
            结束单元格，如 "D10"。

        Returns
        -------
        list[list[Any]]
            二维列表。
        """
        filepath = str(filepath)
        path = ExcelTool._ensure_path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {filepath}")

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            ws = ExcelTool._get_sheet(wb, sheet_name)
            cell_range = ws[f"{start_cell}:{end_cell}"]
        finally:
            wb.close()

        result = []
        for row in cell_range:
            result.append([ExcelTool._cell_to_value(cell) for cell in row])
        return result

    @staticmethod
    def write(
        filepath: str,
        data: list[dict],
        sheet_name: str = "Sheet1",
    ) -> None:
        """
        写入 Excel 文件。

        data 的 key 作为表头（第一行），value 作为数据行。
        如果文件不存在 → 创建新文件。
        如果 sheet 不存在 → 创建新 sheet。
        如果 sheet 已存在 → 清除后重写。

        Parameters
        ----------
        filepath : str
            目标文件路径。
        data : list[dict]
            待写入的数据。
        sheet_name : str
            Sheet 名称，默认 "Sheet1"。
        """
        filepath = str(filepath)
        path = Path(filepath)

        # 打开或创建工作簿
        if path.exists():
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()

        try:
            # 处理 sheet
            if sheet_name in wb.sheetnames:
                # 移除旧 sheet 后重建（保留位置）
                idx = wb.sheetnames.index(sheet_name)
                del wb[sheet_name]
                ws = wb.create_sheet(title=sheet_name, index=idx)
            else:
                # 如果只有默认 Sheet 且为空，重命名它
                if len(wb.sheetnames) == 1 and wb.active.title == "Sheet" and not list(wb.active.iter_rows()):
                    ws = wb.active
                    ws.title = sheet_name
                else:
                    ws = wb.create_sheet(title=sheet_name)

            if not data:
                logger.info("data 为空，仅创建 sheet '%s'（无数据行）", sheet_name)
                wb.save(path)
                return

            # 收集表头（按 data 中 key 的首次出现顺序）
            headers = list(data[0].keys())

            # 写表头
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # 写数据行
            for row_idx, record in enumerate(data, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=record.get(header))

            wb.save(path)
            logger.info("成功写入 %d 行数据到 %s → %s", len(data), path.name, sheet_name)
        finally:
            wb.close()

    @staticmethod
    def get_sheet_names(filepath: str) -> list[str]:
        """返回所有 sheet 名称。"""
        filepath = str(filepath)
        path = ExcelTool._ensure_path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {filepath}")

        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            return wb.sheetnames
        finally:
            wb.close()

    @staticmethod
    def get_info(filepath: str) -> dict:
        """
        返回文件信息。

        Returns
        -------
        dict
            {
                "sheets": ["Sheet1", "Sheet2"],
                "rows": {"Sheet1": 100, "Sheet2": 50},
                "columns": {"Sheet1": 26, "Sheet2": 10},
                "filepath": "D:/data/example.xlsx"
            }
        """
        filepath = str(filepath)
        path = ExcelTool._ensure_path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {filepath}")

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            sheets = wb.sheetnames
            rows_info = {}
            cols_info = {}
            for name in sheets:
                ws = wb[name]
                rows_info[name] = ws.max_row or 0
                cols_info[name] = ws.max_column or 0
        finally:
            wb.close()

        return {
            "sheets": sheets,
            "rows": rows_info,
            "columns": cols_info,
            "filepath": str(path.resolve()),
        }

    @staticmethod
    def create_workbook(
        filepath: str,
        sheets: list[str] | None = None,
    ) -> None:
        """
        创建新的 Excel 文件，可选多个 sheet。

        Parameters
        ----------
        filepath : str
            目标文件路径。
        sheets : list[str] | None
            Sheet 名称列表。None 时创建一个默认 "Sheet"。
        """
        filepath = str(filepath)
        path = Path(filepath)
        if path.exists():
            raise FileExistsError(f"文件已存在，请使用其他路径: {filepath}")

        wb = openpyxl.Workbook()

        try:
            if sheets:
                for i, name in enumerate(sheets):
                    if i == 0:
                        wb.active.title = name
                    else:
                        wb.create_sheet(title=name)
            wb.save(path)
            logger.info("创建 Excel 文件: %s，sheets: %s", path.name, wb.sheetnames)
        finally:
            wb.close()

    @staticmethod
    def copy_sheet(
        filepath: str,
        source_sheet: str,
        target_sheet: str,
    ) -> None:
        """
        复制 sheet。

        在同一个工作簿内将 source_sheet 复制为 target_sheet。
        如果 target_sheet 已存在则先删除。
        """
        filepath = str(filepath)
        path = ExcelTool._ensure_path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {filepath}")

        wb = openpyxl.load_workbook(path)
        try:
            try:
                src_ws = wb[source_sheet]
            except KeyError:
                raise ValueError(
                    f"源 sheet '{source_sheet}' 不存在。可用: {wb.sheetnames}"
                ) from None

            # 如果目标已存在则删除
            if target_sheet in wb.sheetnames:
                del wb[target_sheet]

            copied = wb.copy_worksheet(src_ws)
            copied.title = target_sheet

            wb.save(path)
            logger.info(
                "复制 sheet: '%s' → '%s' (文件: %s)",
                source_sheet,
                target_sheet,
                path.name,
            )
        finally:
            wb.close()

    @staticmethod
    def open_in_wps(filepath: str) -> bool:
        """
        用 xlwings 在 WPS/Excel 中打开文件。

        Returns
        -------
        bool
            True 表示成功打开，False 表示 xlwings 不可用或打开失败。
        """
        filepath = str(filepath)
        path = Path(filepath)
        if not path.exists():
            logger.error("文件不存在，无法用 WPS 打开: %s", filepath)
            return False

        try:
            import xlwings as xw
        except ImportError:
            logger.warning("xlwings 未安装，无法通过 COM 打开 WPS/Excel")
            return False

        try:
            # visible=True 让用户能看到打开的窗口
            app = xw.App(visible=True)
            wb = app.books.open(str(path.resolve()))
            logger.info("已在 WPS/Excel 中打开: %s", path.name)
            return True
        except Exception as exc:
            logger.error("通过 xlwings 打开文件失败: %s", exc)
            return False
